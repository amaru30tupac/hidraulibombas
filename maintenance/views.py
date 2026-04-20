from typing import Dict, List, Tuple
import base64
import os
from uuid import uuid4

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth, TruncWeek
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from clients.models import Client
from users.models import User
from .forms import MaintenanceRecordForm
from .models import InspectionValue, MaintenancePhoto, MaintenanceRecord
STATE_LABELS = {
    "": "—",
    "A": "Ajustadas",
    "C": "Cambio",
    "F": "Falla",
    "P": "Pendiente",
    "D": "Diagnóstico",
    "V": "Verificado / OK / Operando",
    "NA": "No aplica",
}

TABLERO_CONTROL_ROWS = [
    ("variador", "Variador", "state"),
    ("muletillas", "Muletillas", "state"),
    ("piloto", "Piloto", "state"),
    ("contactores", "Contactores", "state"),
    ("presostatos", "Presostatos", "state"),
    ("alternador", "Alternador", "state"),
    ("guarda_motor", "Guarda motor", "state"),
    ("temporizador", "Temporizador", "state"),
    ("rele_termico", "Relé térmico", "state"),
    ("flotador_electrico", "Flotador eléctrico", "state"),
    ("conexiones", "Conexiones", "state"),
    ("amperaje", "Amperaje", "number"),
    ("voltaje", "Voltaje", "number"),
]

HIDROSISTEMA_ROWS = [
    ("precarga", "Precarga"),
    ("flotador_mecanico", "Flotador mecánico"),
    ("valvulas_succion", "Válvulas de succión"),
    ("tanque_hidro", "Tanque hidro"),
    ("cheques", "Cheques"),
    ("presion_linea", "Presión de línea"),
    ("registros", "Registros"),
    ("membrana", "Membrana"),
    ("manometro", "Manómetro"),
    ("cargador_aire", "Cargador de aire"),
    ("tanque_reserva", "Tanque reserva"),
    ("flauta_descarga", "Flauta de descarga"),
]

SISTEMA_EYECTOR_ROWS = [
    ("guarda_motor", "Guarda motor", "state"),
    ("breaker", "Breaker", "state"),
    ("muletilla", "Muletilla", "state"),
    ("contactores", "Contactores", "state"),
    ("flotador_electrico", "Flotador eléctrico", "state"),
    ("conexiones", "Conexiones", "state"),
    ("amperaje", "Amperaje", "number"),
    ("voltaje", "Voltaje", "number"),
    ("sirena", "Sirena", "state"),
    ("rele_termico", "Relé térmico", "state"),
    ("residuos_solidos", "Residuos sólidos", "state"),
]

MOTOR_ROWS = [
    ("rodamientos", "Rodamientos", "state"),
    ("casquillo", "Casquillo", "state"),
    ("empaque", "Empaque", "state"),
    ("ventilador", "Ventilador", "state"),
    ("bornera", "Bornera", "state"),
    ("bobinado", "Bobinado", "state"),
]

BOMBAS_ROWS = [
    ("impulsor", "Impulsor", "state"),
    ("sello_mecanico", "Sello mecánico", "state"),
    ("empaque", "Empaque", "state"),
    ("carcasa", "Carcasa", "state"),
    ("casquillo", "Casquillo", "state"),
]


def build_state_choices() -> List[Tuple[str, str]]:
    return [
        ("", "—"),
        ("A", "A – Ajustadas"),
        ("C", "C – Cambio"),
        ("F", "F – Falla"),
        ("P", "P – Pendiente"),
        ("D", "D – Diagnóstico"),
        ("V", "V – Verificado / OK / Operando"),
        ("NA", "N/A – No aplica"),
    ]


def build_table_rows_for_template(rows_config, positions, section_name, existing_map=None):
    existing_map = existing_map or {}
    built_rows = []

    for row in rows_config:
        key = row[0]
        label = row[1]
        row_type = row[2] if len(row) > 2 else "state"

        cells = []
        for position in positions:
            existing = existing_map.get((section_name, key, position))
            cell_name = f"{section_name}__{key}__{position}"

            value = ""
            numeric_value = ""

            if existing:
                value = existing.value or ""
                numeric_value = "" if existing.numeric_value is None else str(existing.numeric_value)

            cells.append({
                "name": cell_name,
                "type": row_type,
                "value": value,
                "numeric_value": numeric_value,
                "position": position,
            })

        built_rows.append({
            "key": key,
            "label": label,
            "type": row_type,
            "cells": cells,
        })

    return built_rows


def build_hidrosistema_rows_for_template(rows_config, existing_map=None):
    existing_map = existing_map or {}
    built_rows = []

    for key, label in rows_config:
        existing = existing_map.get(("hidrosistema", key, "estado"))
        built_rows.append({
            "key": key,
            "label": label,
            "name": f"hidrosistema__{key}__estado",
            "value": existing.value if existing else "",
        })

    return built_rows


def calculate_general_status_from_post(post_data: Dict[str, str]) -> str:
    values = [value for key, value in post_data.items() if "__" in key and value]

    if "F" in values:
        return "Con fallas"
    if "P" in values:
        return "Pendiente"
    if "D" in values:
        return "En diagnóstico"
    if values:
        return "Revisado"
    return "Sin clasificar"


@transaction.atomic
def save_inspection_values(record: MaintenanceRecord, post_data):
    InspectionValue.objects.filter(maintenance=record).delete()

    for key, value in post_data.items():
        if "__" not in key:
            continue

        section, parameter, position = key.split("__", 2)

        if section not in {
            "tablero_control",
            "hidrosistema",
            "sistema_eyector",
            "motor",
            "bombas",
        }:
            continue

        if value == "":
            continue

        if parameter in {"amperaje", "voltaje"}:
            try:
                numeric_value = float(value)
            except ValueError:
                numeric_value = None

            InspectionValue.objects.create(
                maintenance=record,
                section=section,
                parameter=parameter,
                position=position,
                value="",
                numeric_value=numeric_value,
            )
        else:
            InspectionValue.objects.create(
                maintenance=record,
                section=section,
                parameter=parameter,
                position=position,
                value=value,
                numeric_value=None,
            )


@transaction.atomic
def save_uploaded_photos(record: MaintenanceRecord, files):
    for file in files:
        MaintenancePhoto.objects.create(
            maintenance=record,
            image=file,
        )


def get_existing_values_map(record):
    return {
        (item.section, item.parameter, item.position): item
        for item in record.inspection_values.all()
    }


def build_maintenance_form_context(form, page_title, record=None, existing_values=None):
    existing_values = existing_values or {}

    return {
        "page_title": page_title,
        "form": form,
        "record": record,
        "state_choices": build_state_choices(),
        "tablero_rows": build_table_rows_for_template(
            TABLERO_CONTROL_ROWS, ["b1", "b2", "b3", "b4"], "tablero_control", existing_values
        ),
        "hidrosistema_rows": build_hidrosistema_rows_for_template(HIDROSISTEMA_ROWS, existing_values),
        "eyector_rows": build_table_rows_for_template(
            SISTEMA_EYECTOR_ROWS, ["b1", "b2", "b3", "b4"], "sistema_eyector", existing_values
        ),
        "motor_rows": build_table_rows_for_template(
            MOTOR_ROWS, ["m1", "m2", "m3", "m4"], "motor", existing_values
        ),
        "bombas_rows": build_table_rows_for_template(
            BOMBAS_ROWS, ["b1", "b2", "b3", "b4"], "bombas", existing_values
        ),
    }


def get_filtered_records(request):
    query = request.GET.get("q", "").strip()
    client_id = request.GET.get("client", "").strip()
    technician_id = request.GET.get("technician", "").strip()
    maintenance_type = request.GET.get("maintenance_type", "").strip()
    general_status = request.GET.get("general_status", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    records = MaintenanceRecord.objects.select_related("client", "technician").all()

    if query:
        records = records.filter(
            Q(title__icontains=query) |
            Q(client__name__icontains=query) |
            Q(equipment__icontains=query) |
            Q(tower__icontains=query) |
            Q(number__icontains=query)
        )

    if client_id:
        records = records.filter(client_id=client_id)

    if technician_id:
        records = records.filter(technician_id=technician_id)

    if maintenance_type:
        records = records.filter(maintenance_type=maintenance_type)

    if general_status:
        records = records.filter(general_status=general_status)

    if date_from:
        records = records.filter(date__gte=date_from)

    if date_to:
        records = records.filter(date__lte=date_to)

    return records.order_by("-date", "-created_at")


@login_required
def maintenance_list_view(request):
    records = get_filtered_records(request)
    paginator = Paginator(records, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    clients = Client.objects.filter(is_active=True).order_by("name")
    technicians = User.objects.filter(is_active=True).order_by("first_name", "last_name", "username")

    context = {
        "page_title": "Mantenimientos",
        "records": page_obj,
        "page_obj": page_obj,
        "clients": clients,
        "technicians": technicians,
        "query": request.GET.get("q", "").strip(),
        "selected_client": request.GET.get("client", "").strip(),
        "selected_technician": request.GET.get("technician", "").strip(),
        "selected_type": request.GET.get("maintenance_type", "").strip(),
        "selected_status": request.GET.get("general_status", "").strip(),
        "date_from": request.GET.get("date_from", "").strip(),
        "date_to": request.GET.get("date_to", "").strip(),
    }
    return render(request, "maintenance/maintenance_list.html", context)


@login_required
def maintenance_detail_view(request, pk):
    record = get_object_or_404(
        MaintenanceRecord.objects.select_related("client", "technician"),
        pk=pk,
    )

    inspection_values = record.inspection_values.all()
    grouped = {}
    for item in inspection_values:
        grouped.setdefault(item.section, []).append(item)

    return render(request, "maintenance/maintenance_detail.html", {
        "page_title": f"Detalle #{record.id}",
        "record": record,
        "grouped_values": grouped,
        "state_labels": STATE_LABELS,
    })


@login_required
@transaction.atomic
def maintenance_create_view(request):
    if request.method == "POST":
        form = MaintenanceRecordForm(request.POST, request.FILES)

        if form.is_valid():
            record = form.save(commit=False)
            record.created_by = request.user
            record.technician = request.user
            record.general_status = calculate_general_status_from_post(request.POST)

            if not record.entry_time:
                record.entry_time = timezone.localtime().time().replace(second=0, microsecond=0)

            record.title = f"{record.client.name} - {record.date} - {record.equipment}"
            record.save()

            save_inspection_values(record, request.POST)
            save_uploaded_photos(record, request.FILES.getlist("photos"))

            messages.success(request, "Mantenimiento creado correctamente.")
            return redirect("maintenance_detail", pk=record.pk)
    else:
        initial_data = {
            "entry_time": timezone.localtime().time().replace(second=0, microsecond=0),
        }
        form = MaintenanceRecordForm(initial=initial_data)

    context = build_maintenance_form_context(form=form, page_title="Nuevo mantenimiento")
    return render(request, "maintenance/maintenance_form.html", context)

@login_required
@transaction.atomic
def maintenance_update_view(request, pk):
    record = get_object_or_404(MaintenanceRecord, pk=pk)
    existing_values = get_existing_values_map(record)

    if record.is_closed:
        messages.error(request, "Este mantenimiento está cerrado y no se puede editar.")
        return redirect("maintenance_detail", pk=record.pk)

    if request.method == "POST":
        form = MaintenanceRecordForm(request.POST, request.FILES, instance=record)

        if form.is_valid():
            record = form.save(commit=False)

            if not record.technician:
                record.technician = request.user

            if not record.entry_time:
                record.entry_time = timezone.localtime().time().replace(second=0, microsecond=0)

            record.general_status = calculate_general_status_from_post(request.POST)
            record.title = f"{record.client.name} - {record.date} - {record.equipment}"
            record.save()

            save_inspection_values(record, request.POST)
            save_uploaded_photos(record, request.FILES.getlist("photos"))

            messages.success(request, "Mantenimiento actualizado correctamente.")
            return redirect("maintenance_detail", pk=record.pk)
    else:
        form = MaintenanceRecordForm(instance=record)

    context = build_maintenance_form_context(
        form=form,
        page_title=f"Editar mantenimiento #{record.id}",
        record=record,
        existing_values=existing_values,
    )
    return render(request, "maintenance/maintenance_form.html", context)

@login_required
def maintenance_delete_view(request, pk):
    record = get_object_or_404(MaintenanceRecord, pk=pk)

    if request.method == "POST":
        record.delete()
        messages.success(request, "Mantenimiento eliminado correctamente.")
        return redirect("maintenance_list")

    return render(request, "maintenance/maintenance_confirm_delete.html", {
        "page_title": "Eliminar mantenimiento",
        "record": record,
    })


@login_required
def maintenance_delete_photo_view(request, photo_id):
    photo = get_object_or_404(MaintenancePhoto, pk=photo_id)
    maintenance_id = photo.maintenance_id

    if request.method == "POST":
        if photo.maintenance.is_closed:
            messages.error(request, "No puedes eliminar fotos de un mantenimiento cerrado.")
            return redirect("maintenance_detail", pk=maintenance_id)

        photo.image.delete(save=False)
        photo.delete()
        messages.success(request, "Foto eliminada correctamente.")

    return redirect("maintenance_update", pk=maintenance_id)






def _safe_text(value):
    return "-" if value in [None, ""] else str(value)


def _pdf_styles_v3():
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="PdfHeaderTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        textColor=colors.HexColor("#2d6f9f"),
        alignment=TA_CENTER,
    ))

    styles.add(ParagraphStyle(
        name="PdfSectionTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#2d6f9f"),
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="PdfHeaderSub",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#2d6f9f"),
        alignment=TA_CENTER,
    ))

    styles.add(ParagraphStyle(
        name="PdfBoxText",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.black,
    ))

    styles.add(ParagraphStyle(
        name="PdfMiniCenter",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.black,
        alignment=TA_CENTER,
    ))

    styles.add(ParagraphStyle(
        name="PdfSectionBar",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#2d6f9f"),
        alignment=TA_LEFT,
    ))

    styles.add(ParagraphStyle(
        name="PdfFooter",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#444444"),
        alignment=TA_CENTER,
    ))

    return styles


def _get_logo_path():
    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logo.png")
    return logo_path if os.path.exists(logo_path) else None


def _state_display(value):
    mapping = {
        "A": "A",
        "C": "C",
        "F": "F",
        "P": "P",
        "D": "D",
        "V": "V",
        "NA": "N/A",
        "": "",
        None: "",
    }
    return mapping.get(value, str(value))


def _inspection_map(record):
    return {
        (item.section, item.parameter, item.position): item
        for item in record.inspection_values.all()
    }


def _inspection_value(inspection_map, section, parameter, position):
    item = inspection_map.get((section, parameter, position))
    if not item:
        return ""
    if item.numeric_value is not None:
        value = item.numeric_value
        return str(int(value)) if float(value).is_integer() else str(value)
    return _state_display(item.value)


def _build_header_block(styles, record):
    logo_path = _get_logo_path()

    logo = Paragraph(" ", styles["PdfMiniCenter"])
    if logo_path:
        try:
            logo = Image(logo_path, width=3.6 * cm, height=1.9 * cm)
        except Exception:
            pass

    right_block = [
        Paragraph(f"REGISTRO DE MANTENIMIENTO # {record.pk}", styles["PdfHeaderTitle"]),
        Paragraph("INGENIERÍA HIDRAULIBOMBAS S.A.S", styles["PdfHeaderSub"]),
        Paragraph("NIT: 901.454.224-1", styles["PdfHeaderSub"]),
    ]

    table = Table([[logo, right_block]], colWidths=[4.2 * cm, 12.3 * cm])
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return table


def _build_general_info_box(styles, record):
    left_lines = [
        f"<b>Cliente:</b> {_safe_text(record.client.name if record.client else '')}",
        f"<b>Torre:</b> {_safe_text(record.tower)}",
        f"<b>Número:</b> {_safe_text(record.number)}",
        f"<b>Hora Entrada:</b> {_safe_text(record.entry_time)}",
        f"<b>Hora Salida:</b> {_safe_text(record.exit_time)}",
    ]

    right_lines = [
        f"<b>Tipo de mantenimiento:</b> {_safe_text(record.get_maintenance_type_display())}",
        f"<b>Equipo:</b> {_safe_text(record.equipment)}",
        f"<b>Técnico:</b> {_safe_text(record.technician)}",
        f"<b>Fecha:</b> {_safe_text(record.date)}",
    ]

    table = Table(
        [[
            Paragraph("<br/>".join(left_lines), styles["PdfBoxText"]),
            Paragraph("<br/>".join(right_lines), styles["PdfBoxText"]),
        ]],
        colWidths=[8.2 * cm, 8.3 * cm],
    )
    table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor("#cfd8df")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dbe3e8")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
    ]))
    return table


def _build_top_image_block(styles, title, image_source=None):
    title_p = Paragraph(title, styles["PdfMiniCenter"])

    body = Paragraph(" ", styles["PdfMiniCenter"])
    if image_source:
        try:
            body = Image(image_source, width=2.0 * cm, height=2.0 * cm)
        except Exception:
            body = Paragraph("Sin imagen", styles["PdfMiniCenter"])

    table = Table([[title_p], [body]], colWidths=[8.0 * cm])
    table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("TOPPADDING", (0, 1), (-1, 1), 4),
    ]))
    return table


def _build_section_title(styles, title, width):
    table = Table([[Paragraph(title, styles["PdfSectionBar"])]], colWidths=[width])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#d7e6f1")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#2d6f9f")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def _styled_matrix_table(data, col_widths):
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d5dde4")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("LEADING", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    return table


def _build_tablero_table(inspection_map):
    rows = [
        ("Variador", "variador"),
        ("Muletillas", "muletillas"),
        ("Piloto", "piloto"),
        ("Contactores", "contactores"),
        ("Presostatos", "presostatos"),
        ("Alternador", "alternador"),
        ("Guarda motor", "guarda_motor"),
        ("Temporizador", "temporizador"),
        ("Relé térmico", "rele_termico"),
        ("Flotador eléctrico", "flotador_electrico"),
        ("Conexiones", "conexiones"),
        ("Amperaje", "amperaje"),
        ("Voltaje", "voltaje"),
    ]
    data = [["Parámetro", "B1", "B2", "B3", "B4"]]
    for label, key in rows:
        data.append([
            label,
            _inspection_value(inspection_map, "tablero_control", key, "b1"),
            _inspection_value(inspection_map, "tablero_control", key, "b2"),
            _inspection_value(inspection_map, "tablero_control", key, "b3"),
            _inspection_value(inspection_map, "tablero_control", key, "b4"),
        ])
    return _styled_matrix_table(data, [4.2 * cm, 0.72 * cm, 0.72 * cm, 0.72 * cm, 0.72 * cm])


def _build_hidrosistema_table(inspection_map):
    rows = [
        ("Precarga", "precarga"),
        ("Flotador mecánico", "flotador_mecanico"),
        ("Válvulas de succión", "valvulas_succion"),
        ("Tanque hidro", "tanque_hidro"),
        ("Cheques", "cheques"),
        ("Presión línea", "presion_linea"),
        ("Registros", "registros"),
        ("Membrana", "membrana"),
        ("Manómetro", "manometro"),
        ("Cargador de aire", "cargador_aire"),
        ("Tanque reserva", "tanque_reserva"),
        ("Flauta de descarga", "flauta_descarga"),
    ]
    data = [["Descripción", "Estado"]]
    for label, key in rows:
        data.append([label, _inspection_value(inspection_map, "hidrosistema", key, "estado")])
    return _styled_matrix_table(data, [3.4 * cm, 1.0 * cm])


def _build_eyector_table(inspection_map):
    rows = [
        ("Guarda motor", "guarda_motor"),
        ("Breaker", "breaker"),
        ("Muletilla", "muletilla"),
        ("Contactores", "contactores"),
        ("Flotador eléctrico", "flotador_electrico"),
        ("Conexiones", "conexiones"),
        ("Amperaje", "amperaje"),
        ("Voltaje", "voltaje"),
        ("Sirena", "sirena"),
        ("Relé térmico", "rele_termico"),
        ("Residuos sólidos", "residuos_solidos"),
    ]
    data = [["Parámetro", "B1", "B2", "B3", "B4"]]
    for label, key in rows:
        data.append([
            label,
            _inspection_value(inspection_map, "sistema_eyector", key, "b1"),
            _inspection_value(inspection_map, "sistema_eyector", key, "b2"),
            _inspection_value(inspection_map, "sistema_eyector", key, "b3"),
            _inspection_value(inspection_map, "sistema_eyector", key, "b4"),
        ])
    return _styled_matrix_table(data, [2.5 * cm, 0.5 * cm, 0.5 * cm, 0.5 * cm, 0.5 * cm])


def _build_motor_table(inspection_map):
    rows = [
        ("Rodamientos", "rodamientos"),
        ("Casquillo", "casquillo"),
        ("Empaque", "empaque"),
        ("Ventilador", "ventilador"),
        ("Bornera", "bornera"),
        ("Bobinado", "bobinado"),
    ]
    data = [["Parámetro", "M1", "M2", "M3", "M4"]]
    for label, key in rows:
        data.append([
            label,
            _inspection_value(inspection_map, "motor", key, "m1"),
            _inspection_value(inspection_map, "motor", key, "m2"),
            _inspection_value(inspection_map, "motor", key, "m3"),
            _inspection_value(inspection_map, "motor", key, "m4"),
        ])
    return _styled_matrix_table(data, [4.2 * cm, 0.72 * cm, 0.72 * cm, 0.72 * cm, 0.72 * cm])


def _build_bombas_table(inspection_map):
    rows = [
        ("Impulsor", "impulsor"),
        ("Sello mecánico", "sello_mecanico"),
        ("Empaque", "empaque"),
        ("Carcasa", "carcasa"),
        ("Casquillo", "casquillo"),
    ]
    data = [["Parámetro", "B1", "B2", "B3", "B4"]]
    for label, key in rows:
        data.append([
            label,
            _inspection_value(inspection_map, "bombas", key, "b1"),
            _inspection_value(inspection_map, "bombas", key, "b2"),
            _inspection_value(inspection_map, "bombas", key, "b3"),
            _inspection_value(inspection_map, "bombas", key, "b4"),
        ])
    return _styled_matrix_table(data, [4.0 * cm, 0.65 * cm, 0.65 * cm, 0.65 * cm, 0.65 * cm])


def _boxed_text_area(styles, title, content, width):
    title_bar = _build_section_title(styles, title, width)
    body = Table([[Paragraph(_safe_text(content), styles["PdfBoxText"])]], colWidths=[width])
    body.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#d5dde4")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("MINROWHEIGHT", (0, 0), (-1, -1), 1.1 * cm),
    ]))
    return [title_bar, body]


def _footer_table(styles):
    footer_text = "Carrera 96ª N°65 - 52  |  ingehidraulibombas@gmail.com  |  Tel.: 3115898023 - 6016496010"
    table = Table([[Paragraph(footer_text, styles["PdfFooter"])]], colWidths=[17.0 * cm])
    table.setStyle(TableStyle([
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table






















@login_required
def maintenance_export_detail_pdf_view(request, pk):
    record = get_object_or_404(
        MaintenanceRecord.objects.select_related("client", "technician"),
        pk=pk,
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="mantenimiento_{record.pk}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=0.8 * cm,
        leftMargin=0.8 * cm,
        topMargin=0.6 * cm,
        bottomMargin=0.5 * cm,
    )

    styles = _pdf_styles_v3()
    elements = []
    inspection_map = _inspection_map(record)

    elements.append(_build_header_block(styles, record))
    elements.append(Spacer(1, 0.22 * cm))
    elements.append(_build_general_info_box(styles, record))
    elements.append(Spacer(1, 0.32 * cm))

    client_img_path = record.client_signature.path if getattr(record, "client_signature", None) else None
    tech_img_path = record.technician.photo.path if getattr(record.technician, "photo", None) else None

    top_images = Table(
        [[
            _build_top_image_block(styles, "Foto del Cliente", client_img_path),
            _build_top_image_block(styles, "Foto del Técnico", tech_img_path),
        ]],
        colWidths=[8.25 * cm, 8.25 * cm],
    )
    top_images.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(top_images)
    elements.append(Spacer(1, 0.3 * cm))

    tablero_block = [
        _build_section_title(styles, "Tablero de Control", 7.35 * cm),
        _build_tablero_table(inspection_map),
    ]
    hidrosistema_block = [
        _build_section_title(styles, "Hidrosistema", 4.45 * cm),
        _build_hidrosistema_table(inspection_map),
    ]
    eyector_block = [
        _build_section_title(styles, "Sistema de Eyector", 4.75 * cm),
        _build_eyector_table(inspection_map),
    ]

    top_tables = Table(
        [[tablero_block, hidrosistema_block, eyector_block]],
        colWidths=[7.35 * cm, 4.45 * cm, 4.75 * cm],
    )
    top_tables.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(top_tables)
    elements.append(Spacer(1, 0.22 * cm))

    bottom_tables = Table(
        [[
            [_build_section_title(styles, "Motor", 8.0 * cm), _build_motor_table(inspection_map)],
            [_build_section_title(styles, "Bombas", 8.0 * cm), _build_bombas_table(inspection_map)],
        ]],
        colWidths=[8.0 * cm, 8.0 * cm],
    )
    bottom_tables.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(bottom_tables)
    elements.append(Spacer(1, 0.18 * cm))

    elements.append(Paragraph(
        "<b>Estados:</b> A = Aceptable, C = Crítico, F = Fallo, P = Preventivo, D = Desgaste, N/A = No aplica",
        styles["PdfBoxText"]
    ))
    elements.append(Spacer(1, 0.18 * cm))

    notes_table = Table(
        [[
            _boxed_text_area(styles, "Partes para Cambio", record.parts_for_change, 7.95 * cm),
            _boxed_text_area(styles, "Observaciones", record.observations, 7.95 * cm),
        ]],
        colWidths=[8.0 * cm, 8.0 * cm],
    )
    notes_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(notes_table)
    elements.append(Spacer(1, 0.16 * cm))

    approval_title = Paragraph("Firma de aprobación", styles["PdfMiniCenter"])
    approval_name = Paragraph(
        _safe_text(record.client.name if record.client else record.technician),
        styles["PdfMiniCenter"]
    )

    approval_signature = Paragraph(" ", styles["PdfMiniCenter"])
    if getattr(record, "client_signature", None):
        try:
            approval_signature = Image(record.client_signature.path, width=3.0 * cm, height=1.3 * cm)
        except Exception:
            pass

    approval_table = Table(
        [[approval_title], [approval_signature], [approval_name]],
        colWidths=[16.5 * cm],
    )
    approval_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    elements.append(approval_table)
    elements.append(Spacer(1, 0.15 * cm))

    elements.append(_footer_table(styles))
    doc.build(elements)
    return response


@login_required
def maintenance_export_list_pdf_view(request):
    records = get_filtered_records(request)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="mantenimientos_filtrados.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1 * cm,
    )

    story = _build_pdf_header_story(
        "Listado de mantenimientos",
        "Exportación filtrada desde el sistema.",
    )

    data = [["ID", "Cliente", "Equipo", "Tipo", "Fecha", "Técnico", "Estado", "Proceso"]]

    for record in records:
        data.append([
            str(record.pk),
            _safe_text(record.client.name),
            _safe_text(record.equipment),
            _safe_text(record.get_maintenance_type_display()),
            _safe_text(record.date),
            _safe_text(record.technician),
            _safe_text(record.general_status),
            _safe_text(record.get_process_status_display()),
        ])

    table = Table(
        data,
        colWidths=[1.3 * cm, 6 * cm, 5 * cm, 4 * cm, 3 * cm, 5 * cm, 4 * cm, 4 * cm]
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    story.append(table)
    doc.build(story)
    return response

@login_required
def maintenance_export_excel_view(request):
    records = get_filtered_records(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"

    headers = [
        "ID", "Cliente", "Equipo", "Tipo", "Fecha",
        "Técnico", "Estado", "Proceso", "Torre", "Número", "Sede", "Cerrado"
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for r in records:
        ws.append([
            r.pk,
            r.client.name,
            r.equipment,
            r.get_maintenance_type_display(),
            str(r.date),
            str(r.technician or ""),
            r.general_status,
            r.get_process_status_display(),
            r.tower,
            r.number,
            r.site,
            "Sí" if r.is_closed else "No",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="mantenimientos.xlsx"'

    wb.save(response)
    return response


def _save_signature_from_data_url(data_url: str, prefix: str) -> ContentFile:
    format_part, imgstr = data_url.split(";base64,")
    ext = format_part.split("/")[-1]
    file_name = f"{prefix}_{uuid4().hex}.{ext}"
    return ContentFile(base64.b64decode(imgstr), name=file_name)


@login_required
@transaction.atomic
def maintenance_sign_technician_view(request, pk):
    record = get_object_or_404(MaintenanceRecord, pk=pk)

    if request.method == "POST":
        signature_data = request.POST.get("signature_data", "").strip()
        if not signature_data:
            messages.error(request, "Debes firmar antes de guardar.")
            return redirect("maintenance_detail", pk=pk)

        file_content = _save_signature_from_data_url(signature_data, "tech_signature")
        if record.technician_signature:
            record.technician_signature.delete(save=False)

        record.technician_signature.save(file_content.name, file_content, save=False)
        record.signed_by_technician_at = timezone.now()

        if record.process_status == "borrador":
            record.process_status = "en_proceso"

        record.save()
        messages.success(request, "Firma del técnico guardada correctamente.")

    return redirect("maintenance_detail", pk=pk)


@login_required
@transaction.atomic
def maintenance_sign_client_view(request, pk):
    record = get_object_or_404(MaintenanceRecord, pk=pk)

    if request.method == "POST":
        signature_data = request.POST.get("signature_data", "").strip()
        if not signature_data:
            messages.error(request, "Debes firmar antes de guardar.")
            return redirect("maintenance_detail", pk=pk)

        file_content = _save_signature_from_data_url(signature_data, "client_signature")
        if record.client_signature:
            record.client_signature.delete(save=False)

        record.client_signature.save(file_content.name, file_content, save=False)
        record.signed_by_client_at = timezone.now()

        if record.technician_signature and record.client_signature:
            record.is_closed = True
            record.process_status = "cerrado"

        record.save()
        messages.success(request, "Firma del cliente guardada correctamente.")

    return redirect("maintenance_detail", pk=pk)